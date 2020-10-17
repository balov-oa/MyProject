import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows

from excel_styles import (
    base_style,
    headers_style,
    text_wrap_style,
)
from sql_connector import get_sqlalchemy_engine

ENGINE = get_sqlalchemy_engine()
BASE_PATH = Path(__file__).parent.parent


def get_data_for_make_excel() -> pd.DataFrame:
    query = """SELECT 
                     Apartments.Url_Link,
                     CAST(Apartments.Date_Add as date) AS Date_Add,
                     CAST(Apartments.Date_Expiration as date) AS Date_Expiration,
                     Apartments.District,
                     Apartments.Address,
                     Apartments.Year_Building,
                     Apartments.Material,
                     Apartments.Floor_Numbers_Of_Floors,
                     Apartments.Square_Total,
                     Apartments.Apartment_Condition,
                     Apartments.Price / 1000 AS Price,
                     Predictions.Predict,
                     Predictions.Error
               FROM Apartment_Tomsk.dbo.Predictions AS Predictions
               INNER JOIN Apartment_Tomsk.dbo.Apartments AS Apartments
                   ON Predictions.Apartment_Key = Apartments.Apartment_Key
               WHERE Error <=0
               ORDER BY Date_Add DESC, Error"""

    data = pd.read_sql_query(query, ENGINE)
    map_rename = {
        'Url_Link': 'Ссылка',
        'Date_Add': 'Дата добавления',
        'Date_Expiration': 'Дата истечения',
        'District': 'Район',
        'Address': 'Адрес',
        'Year_Building': 'Год постройки',
        'Material': 'Материал',
        'Floor_Numbers_Of_Floors': 'Этаж/этажность',
        'Square_Total': 'Площадь',
        'Apartment_Condition': 'Состояние',
        'Price': 'Цена',
        'Predict': 'Прогноз',
        'Error': 'Ошибка',
    }
    data.rename(columns=map_rename, inplace=True)
    data.set_index('Дата добавления', inplace=True, drop=True)
    data.index = pd.to_datetime(data.index)
    data_out = data.groupby(pd.Grouper(freq='D')).apply(lambda x: x.sort_values('Ошибка').iloc[0:10])
    data_out.reset_index(level=0, drop=True, inplace=True)
    data_out.reset_index(inplace=True)
    data_out.sort_values(['Дата добавления', 'Ошибка'], ascending=[False, True], inplace=True)
    data_out['Дата добавления'] = data_out['Дата добавления'].astype(str)
    data_out['Дата истечения'] = data_out['Дата истечения'].astype(str)
    return data_out


def make_excel(data_out: pd.DataFrame) -> None:
    wb = Workbook()
    wb.add_named_style(headers_style)
    wb.add_named_style(base_style)
    wb.add_named_style(text_wrap_style)
    ws = wb.active
    ws.auto_filter.ref = 'A1:M1'

    # Add data
    for r in dataframe_to_rows(data_out, index=False, header=True):
        ws.append(r)

    # Set Base style
    for columns in ws['A:M']:
        for cell in columns[1:]:
            cell.style = base_style

    # Set style for Headers
    cells = ws['A1:M1']
    for cell in cells[0]:
        cell.style = headers_style

    # Set style for Hyperlink
    for cell in ws['B:B'][1:]:
        cell.hyperlink = cell.value
        cell.style = 'Hyperlink'

    for column in ['D', 'E', 'J']:
        for cell in ws[f'{column}:{column}'][1:]:
            cell.style = text_wrap_style

    # Set width for columns
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
    for column in columns:
        ws.column_dimensions[column].width = 13
    for row in range(len(ws['A'])):
        ws.row_dimensions[row].height = 25.5

    start = 2
    list_to_merge = []
    for i in data_out['Дата добавления'].value_counts().sort_index(ascending=False).values:
        list_to_merge.append([start, start + i - 1])
        start += i

    flag = 2
    for i, j in list_to_merge:
        ws.merge_cells(f'A{i}:A{j}')
        if flag % 2 == 0:
            ws[f'A{i}'].fill = headers_style.fill
        ws[f'A{i}'].font = headers_style.font
        ws[f'A{i}'].alignment = headers_style.alignment
        flag += 1
        ws.conditional_formatting.add(f'M{i}:M{j}',
                                      ColorScaleRule(start_type='max', end_type='min', start_color='CAFFBF',
                                                     end_color='FFD6A5')
                                      )

    wb.save(BASE_PATH.parent.joinpath('OneDrive').joinpath('report.xlsx'))


if __name__ == '__main__':
    log_file = BASE_PATH.joinpath('logs').joinpath('make_excel.txt')
    logging.basicConfig(
        format='[%(asctime)s] -- %(levelname).3s -- %(message)s',
        datefmt='%Y.%m.%d %H:%M:%S',
        level=logging.DEBUG,
        filename=log_file)

    logging.info('Make excel')
    try:
        make_excel(get_data_for_make_excel())
    except Exception as e:
        logging.exception(e)
